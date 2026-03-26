import argparse
import csv
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook
import requests

BASE_URL = "https://api.weather.gc.ca/collections/climate-stations/items/"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate ICAO codes from Excel against Weather Canada station data."
    )
    parser.add_argument("--excel", required=True, help="Path to Excel file containing ICAO codes")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    parser.add_argument(
        "--column",
        default=None,
        help="Column name containing ICAO codes (default: auto-detect common names)",
    )
    parser.add_argument(
        "--output",
        default="icao_validation_results.csv",
        help="Output CSV path for validation results",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=1000,
        help="API page size when downloading stations (default: 1000)",
    )
    return parser.parse_args()


def normalize_icao(value: str) -> str:
    return str(value).strip().upper()


def choose_column(columns: list[str], user_column: str | None) -> str:
    if user_column:
        if user_column not in columns:
            raise ValueError(f"Column '{user_column}' not found. Available: {columns}")
        return user_column

    candidates = [
        "ICAO",
        "Icao",
        "icao",
        "TC_IDENTIFIER",
        "TC Identifier",
        "Code",
        "CODE",
    ]
    for candidate in candidates:
        if candidate in columns:
            return candidate

    return columns[0]


def load_icao_codes(excel_path: Path, sheet: str | None, column: str | None) -> list[str]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        if not any(header):
            return []

        code_column = choose_column(header, column)
        code_idx = header.index(code_column)

        values = []
        for row in rows[1:]:
            if code_idx >= len(row):
                continue
            raw = row[code_idx]
            if raw is None:
                continue
            code = normalize_icao(raw)
            if code:
                values.append(code)
    finally:
        wb.close()

    unique_codes: list[str] = []
    seen: set[str] = set()
    for code in values:
        if code not in seen:
            seen.add(code)
            unique_codes.append(code)

    print(f"Loaded {len(unique_codes)} unique ICAO codes from column '{code_column}'.")
    return unique_codes


def fetch_all_stations(limit: int) -> list[dict]:
    params = {"limit": limit, "offset": 0, "f": "json"}
    all_items: list[dict] = []

    while True:
        response = requests.get(BASE_URL, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        items = data.get("features") or data.get("items") or []
        if not items:
            break

        all_items.extend(items)
        print(f"Fetched stations: {len(all_items)}")

        if len(items) < params["limit"]:
            break
        params["offset"] += params["limit"]

    return all_items


def extract_raw_metar_line(text: str) -> str | None:
    lines = [line.strip() for line in str(text or "").splitlines() if line.strip()]
    for line in lines:
        if re.search(r"\b(?:SPECI|METAR)\s+[A-Z]{4}\s+\d{6}Z\b", line):
            return line
    for line in lines:
        if re.search(r"\b[A-Z]{4}\s+\d{6}Z\b", line):
            return line
    for line in lines:
        if line.startswith("METAR ") or line.startswith("SPECI "):
            return line
    return None


def fetch_metar_for_icao(icao: str, timeout: int = 8) -> tuple[bool, str, str, str]:
    clean = normalize_icao(icao)

    aviation_url = f"https://aviationweather.gov/api/data/metar?ids={clean}&format=json"
    aviation_allorigins = f"https://api.allorigins.win/raw?url={quote(aviation_url, safe='')}"
    aviation_codetabs = f"https://api.codetabs.com/v1/proxy?quest={quote(aviation_url, safe='')}"

    noaa_txt = f"https://tgftp.nws.noaa.gov/data/observations/metar/stations/{clean}.TXT"
    noaa_allorigins = f"https://api.allorigins.win/raw?url={quote(noaa_txt, safe='')}"
    noaa_codetabs = f"https://api.codetabs.com/v1/proxy?quest={quote(noaa_txt, safe='')}"

    attempts = [
        ("AviationWeather JSON", aviation_url, "json"),
        ("AviationWeather JSON via AllOrigins", aviation_allorigins, "json"),
        ("AviationWeather JSON via CodeTabs", aviation_codetabs, "json"),
        ("NOAA direct", noaa_txt, "text"),
        ("NOAA via AllOrigins", noaa_allorigins, "text"),
        ("NOAA via CodeTabs", noaa_codetabs, "text"),
    ]

    errors: list[str] = []

    for label, url, mode in attempts:
        try:
            response = requests.get(url, timeout=timeout)
            if response.status_code != 200:
                errors.append(f"{label}: HTTP {response.status_code}")
                continue

            if mode == "json":
                data = response.json()
                if isinstance(data, list) and data:
                    first = data[0]
                    raw = str(first.get("rawOb") or first.get("raw_text") or "").strip()
                    if raw:
                        return True, label, url, raw
                errors.append(f"{label}: empty response")
                continue

            raw_line = extract_raw_metar_line(response.text)
            if raw_line:
                return True, label, url, raw_line
            errors.append(f"{label}: no METAR line")
        except Exception as exc:
            errors.append(f"{label}: {exc}")

    return False, "", "", "; ".join(errors)


def build_station_index(stations: list[dict]) -> dict[str, list[dict]]:
    index: dict[str, list[dict]] = {}
    for feature in stations:
        props = feature.get("properties", {})
        code = normalize_icao(props.get("TC_IDENTIFIER", ""))
        if not code:
            continue
        index.setdefault(code, []).append(props)
    return index


def validate_codes(codes: list[str], station_index: dict[str, list[dict]]) -> list[dict]:
    results: list[dict] = []

    for idx, code in enumerate(codes, start=1):
        matches = station_index.get(code, [])
        metar_ok, metar_source, metar_url, metar_detail = fetch_metar_for_icao(code)
        print(f"Checked {idx}/{len(codes)}: {code} -> {'METAR_OK' if metar_ok else 'NO_METAR'}")

        if matches:
            first = matches[0]
            results.append(
                {
                    "ICAO": code,
                    "station_status": "FOUND",
                    "station_match_count": len(matches),
                    "station_name": first.get("STATION_NAME", ""),
                    "province": first.get("ENG_PROV_NAME", ""),
                    "country": first.get("COUNTRY", ""),
                    "tc_identifier": first.get("TC_IDENTIFIER", ""),
                    "climate_identifier": first.get("CLIMATE_IDENTIFIER", ""),
                    "wmo_identifier": first.get("WMO_IDENTIFIER", ""),
                    "has_hourly_data": first.get("HAS_HOURLY_DATA", ""),
                    "metar_status": "FOUND" if metar_ok else "NOT_FOUND",
                    "metar_source": metar_source,
                    "metar_source_url": metar_url,
                    "metar_detail": metar_detail,
                }
            )
        else:
            results.append(
                {
                    "ICAO": code,
                    "station_status": "NOT_FOUND",
                    "station_match_count": 0,
                    "station_name": "",
                    "province": "",
                    "country": "",
                    "tc_identifier": "",
                    "climate_identifier": "",
                    "wmo_identifier": "",
                    "has_hourly_data": "",
                    "metar_status": "FOUND" if metar_ok else "NOT_FOUND",
                    "metar_source": metar_source,
                    "metar_source_url": metar_url,
                    "metar_detail": metar_detail,
                }
            )

    return results


def write_results_csv(results: list[dict], output_path: Path) -> None:
    if not results:
        print("No validation results to write.")
        return

    fieldnames = list(results[0].keys())
    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)


def main() -> None:
    args = parse_args()

    excel_path = Path(args.excel)
    output_path = Path(args.output)

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    icao_codes = load_icao_codes(excel_path, args.sheet, args.column)
    if not icao_codes:
        print("No ICAO codes found in the selected sheet/column.")
        return

    stations = fetch_all_stations(limit=args.limit)
    station_index = build_station_index(stations)

    results = validate_codes(icao_codes, station_index)
    write_results_csv(results, output_path)

    metar_found = sum(1 for row in results if row["metar_status"] == "FOUND")
    metar_missing = len(results) - metar_found
    station_found = sum(1 for row in results if row["station_status"] == "FOUND")
    station_missing = len(results) - station_found
    print(f"METAR validation complete: {metar_found} found, {metar_missing} not found.")
    print(f"Station index coverage: {station_found} found, {station_missing} not found.")
    print(f"Results written to: {output_path}")


if __name__ == "__main__":
    main()
